import sqlite3
import datetime
import openpyxl
from openpyxl import load_workbook
import re
import os

def huaBSformat(x):  
    if x[-2] == '_':
        x = x[:-2]
    nums = re.findall(r'\d+', x) 
    y = ''.join(nums)
    return int(y)
    
def NSNBSformat(x, n):
    bsNumber = x[-n:]
    return int(bsNumber)


workDir = os.getcwd()    
con = sqlite3.connect(r'link to DB')
cursorObject = con.cursor()
n = 0
cursorObject.execute('SELECT * FROM table in tb')
for i in range(8):
    rowExcel = cursorObject.fetchone()
    if rowExcel[1].startswith('MRB') and huaBSformat(rowExcel[0]) != NSNBSformat(rowExcel[1], 5):
        toBS = huaBSformat(rowExcel[0])
        fromBS = NSNBSformat(rowExcel[1], 5)
        SN = rowExcel[6]
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(len(rowExcel)):
            ws.cell(row = 1, column = i+1).value = rowExcel[i]
            try:
                os.mkdir(workDir + '\\' + f'from {fromBS} to {toBS}')
                os.chdir(workDir + '\\' + f'from {fromBS} to {toBS}')
            except:
                os.chdir(workDir + '\\' + f'from {fromBS} to {toBS}')
            wb.save(f'{SN} from {fromBS} to {toBS}.xlsx')
            wb.close()
con.close()


